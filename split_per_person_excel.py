import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# === KONFIGURASJON ===
INPUT_CSV = "cristin_publikasjoner_kategoriadaptiv.csv"
OUTPUT_DIR = "excel_per_person"

kategori_valg = ["Basic or Discovery", "Applied or Integration", "Teaching and Learning", "N/A"]
type_valg = ["Peer reviewed journal articles", "Additional peer- or editorial reviewed ICs", "All other ICs"]

hjelpetekst = (
    "Velg kategorisering av bidraget, prøv å klassifisere så mange som mulig, bruk kommentarfeltet hvis ikke mulig. "
    "Legg også til en kommentar om det er noe her som er spesielt relevant for HH-s mission statement eller som kan sies å ha hatt impact (målbare resultater ut over siteringer, eks. endring i policy etc.).\n\n"
    "Definisjon av kategoriene:\n"
    "• Basic or Discovery Scholarship is directed toward increasing the knowledge base and the development of theory.\n"
    "• Applied or Integrative/Application Scholarship draws from basic research and uses accumulated theories, knowledge, methods, and techniques to solve real-world problems and/or issues associated with practice.\n"
    "• Teaching and Learning Scholarship explores the theory and methods of teaching and advances new understandings, insights, content, and methods that impact learning behavior."
)

os.makedirs(OUTPUT_DIR, exist_ok=True)

# === Les CSV ===
df = pd.read_csv(INPUT_CSV)

# Legg til tomme kolonner
df["Kategori"] = ""
df["Type"] = ""
df["Kommentar"] = ""
df["Awards etc."] = ""

# Reorganiser kolonnene og fjern uønskede
kolonner = list(df.columns)
kolonner = [k for k in kolonner if k not in ["Cristin-ID", "kanal-kilde"]]

if "Kategori" in kolonner and "Type" in kolonner:
    kolonner.remove("Kategori")
    type_index = kolonner.index("Type")
    kolonner.insert(type_index, "Kategori")

df = df[kolonner]

# === Gruppér per person og lag Excel-filer ===
for navn, gruppe in df.groupby("Navn"):
    navn_deler = navn.strip().split()
    etternavn = navn_deler[-1] if len(navn_deler) >= 2 else navn
    fornavn = " ".join(navn_deler[:-1]) if len(navn_deler) >= 2 else ""
    safe_navn = f"{etternavn}, {fornavn}".strip().replace("/", "_").replace("\\", "_")
    filnavn = f"{OUTPUT_DIR}/publikasjoner - {safe_navn}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Publikasjoner"

    # === 1: Hjelpetekst
    ws.merge_cells(start_row=1, start_column=1, end_row=9, end_column=len(gruppe.columns))
    cell = ws.cell(row=1, column=1)
    cell.value = hjelpetekst
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    # === 2: Skriv overskrifter og data fra rad 11
    data_start_row = 11
    for col_idx, col_name in enumerate(gruppe.columns, start=1):
        ws.cell(row=data_start_row, column=col_idx).value = str(col_name)

    for row_idx, row in enumerate(gruppe.itertuples(index=False), start=data_start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    max_row = ws.max_row
    max_col = ws.max_column

    # === 3: Excel-tabell
    col_letter_end = get_column_letter(max_col)
    tab_ref = f"A{data_start_row}:{col_letter_end}{max_row}"
    table = Table(displayName="Publikasjoner", ref=tab_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # === 4: Nedtrekksmenyer
    headers = [ws.cell(row=data_start_row, column=c).value for c in range(1, max_col + 1)]
    for col in range(1, max_col + 1):
        header = headers[col - 1]
        col_letter = get_column_letter(col)
        if header == "Kategori":
            dv = DataValidation(type="list", formula1=f'"{",".join(kategori_valg)}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{data_start_row + 1}:{col_letter}{max_row}")
        elif header == "Type":
            dv = DataValidation(type="list", formula1=f'"{",".join(type_valg)}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{data_start_row + 1}:{col_letter}{max_row}")

    # === 5: Skjul "Cristin Resultat-ID" hvis finnes
    if "Cristin Resultat-ID" in headers:
        col_idx = headers.index("Cristin Resultat-ID") + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # === 6: Summeringstabell
    sum_start_row = max_row + 2
    ws.cell(row=sum_start_row, column=1).value = "Oppsummering (automatisk)"

    if "Kategori" in headers:
        kat_col_idx = headers.index("Kategori") + 1
        for i, valg in enumerate(kategori_valg):
            ws.cell(row=sum_start_row + 2 + i, column=1).value = f"Antall '{valg}'"
            formel = f'=COUNTIF({get_column_letter(kat_col_idx)}{data_start_row + 1}:{get_column_letter(kat_col_idx)}{max_row},"{valg}")'
            ws.cell(row=sum_start_row + 2 + i, column=2).value = formel
    else:
        ws.cell(row=sum_start_row + 2, column=1).value = "⚠️ Kategori-kolonne ikke funnet"

    if "Type" in headers:
        typ_col_idx = headers.index("Type") + 1
        for i, valg in enumerate(type_valg):
            ws.cell(row=sum_start_row + 8 + i, column=1).value = f"Antall '{valg}'"
            formel = f'=COUNTIF({get_column_letter(typ_col_idx)}{data_start_row + 1}:{get_column_letter(typ_col_idx)}{max_row},"{valg}")'
            ws.cell(row=sum_start_row + 8 + i, column=2).value = formel
    else:
        ws.cell(row=sum_start_row + 8, column=1).value = "⚠️ Type-kolonne ikke funnet"

    wb.save(filnavn)
    print(f"✅ Lagret: {filnavn}")

print("🎯 Ferdig!")

